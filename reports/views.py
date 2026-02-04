"""
Reports views - PDF and Excel exports
"""
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta
import json

from inventory.models import Product, InventoryItem, Category, StockMovement
from transactions.models import EntryVoucher, ExitVoucher, ReturnVoucher, DisposalVoucher


def get_tenant_queryset(request, model):
    """Get queryset filtered by tenant"""
    if request.user.is_super_admin:
        return model.objects.all()
    return model.objects.filter(tenant=request.user.tenant)


@login_required
def reports_index(request):
    """صفحة التقارير الرئيسية"""
    return render(request, 'reports/index.html')


@login_required
def inventory_report(request):
    """تقرير جرد المخزون"""
    items = get_tenant_queryset(request, InventoryItem).select_related('product', 'assigned_to')
    
    # Filters
    status = request.GET.get('status')
    if status:
        items = items.filter(status=status)
    
    category_id = request.GET.get('category')
    if category_id:
        items = items.filter(product__category_id=category_id)
    
    # Summary
    summary = {
        'total': items.count(),
        'available': items.filter(status='available').count(),
        'assigned': items.filter(status='assigned').count(),
        'disposed': items.filter(status='disposed').count(),
        'total_value': items.aggregate(total=Sum('purchase_price'))['total'] or 0,
    }
    
    categories = get_tenant_queryset(request, Category)
    
    context = {
        'items': items,
        'summary': summary,
        'categories': categories,
        'status_choices': InventoryItem.STATUS_CHOICES,
    }
    return render(request, 'reports/inventory_report.html', context)


@login_required
def inventory_report_pdf(request):
    """تقرير جرد المخزون - PDF"""
    items = get_tenant_queryset(request, InventoryItem).select_related('product', 'assigned_to')
    
    status = request.GET.get('status')
    if status:
        items = items.filter(status=status)
    
    summary = {
        'total': items.count(),
        'total_value': items.aggregate(total=Sum('purchase_price'))['total'] or 0,
    }
    
    html = render_to_string('reports/inventory_report_pdf.html', {
        'items': items,
        'summary': summary,
        'tenant': request.user.tenant,
        'date': timezone.now(),
    })
    
    try:
        from weasyprint import HTML
        pdf = HTML(string=html).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response
    except ImportError:
        return HttpResponse("WeasyPrint غير مثبت. يرجى تثبيته لتوليد PDF.", status=500)


@login_required
def inventory_report_excel(request):
    """تقرير جرد المخزون - Excel"""
    items = get_tenant_queryset(request, InventoryItem).select_related('product', 'assigned_to')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الجرد"
        
        # Headers
        headers = ['رقم الجرد', 'الرقم التسلسلي', 'المنتج', 'الصنف', 'الحالة', 'حالة المادة', 'المصلحة', 'سعر الشراء']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item.inventory_number)
            ws.cell(row=row, column=2, value=item.serial_number)
            ws.cell(row=row, column=3, value=item.product.name)
            ws.cell(row=row, column=4, value=item.product.category.name if item.product.category else '')
            ws.cell(row=row, column=5, value=item.get_status_display())
            ws.cell(row=row, column=6, value=item.get_condition_display())
            ws.cell(row=row, column=7, value=item.assigned_to.name if item.assigned_to else '')
            ws.cell(row=row, column=8, value=float(item.purchase_price))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        return HttpResponse("openpyxl غير مثبت. يرجى تثبيته لتوليد Excel.", status=500)


@login_required
def movements_report(request):
    """تقرير حركة المواد"""
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    entries = get_tenant_queryset(request, EntryVoucher).select_related('supplier')
    exits = get_tenant_queryset(request, ExitVoucher).select_related('department')
    returns = get_tenant_queryset(request, ReturnVoucher).select_related('department')
    
    if date_from:
        entries = entries.filter(date__gte=date_from)
        exits = exits.filter(date__gte=date_from)
        returns = returns.filter(date__gte=date_from)
    
    if date_to:
        entries = entries.filter(date__lte=date_to)
        exits = exits.filter(date__lte=date_to)
        returns = returns.filter(date__lte=date_to)
    
    summary = {
        'entries_count': entries.count(),
        'exits_count': exits.count(),
        'returns_count': returns.count(),
    }
    
    context = {
        'entries': entries[:50],
        'exits': exits[:50],
        'returns': returns[:50],
        'summary': summary,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/movements_report.html', context)


@login_required
def disposed_report(request):
    """تقرير المواد التالفة"""
    disposals = get_tenant_queryset(request, DisposalVoucher).prefetch_related('items')
    disposed_items = get_tenant_queryset(request, InventoryItem).filter(status='disposed')
    
    summary = {
        'disposal_vouchers': disposals.count(),
        'disposed_items': disposed_items.count(),
        'total_value': disposed_items.aggregate(total=Sum('purchase_price'))['total'] or 0,
    }
    
    context = {
        'disposals': disposals[:50],
        'disposed_items': disposed_items[:100],
        'summary': summary,
    }
    return render(request, 'reports/disposed_report.html', context)


@login_required
def statistics_api(request):
    """API للإحصائيات - للرسوم البيانية"""
    user = request.user
    tenant = user.tenant
    
    if user.is_super_admin:
        items = InventoryItem.objects.all()
        products = Product.objects.all()
    else:
        items = InventoryItem.objects.filter(tenant=tenant)
        products = Product.objects.filter(tenant=tenant)
    
    # Assets by status
    assets_by_status = list(items.values('status').annotate(count=Count('id')))
    
    # Assets by category
    assets_by_category = list(
        items.values('product__category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    
    # Monthly movements (last 6 months)
    six_months_ago = timezone.now().date() - timedelta(days=180)
    
    monthly_entries = list(
        EntryVoucher.objects.filter(
            tenant=tenant if not user.is_super_admin else tenant,
            date__gte=six_months_ago
        ).extra(select={'month': "strftime('%%Y-%%m', date)"})
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    data = {
        'assets_by_status': assets_by_status,
        'assets_by_category': assets_by_category,
        'monthly_entries': monthly_entries,
    }
    
    return JsonResponse(data)


@login_required
def voucher_pdf(request, voucher_type, pk):
    """طباعة وصل - PDF"""
    models_map = {
        'entry': EntryVoucher,
        'exit': ExitVoucher,
        'return': ReturnVoucher,
        'disposal': DisposalVoucher,
    }
    
    templates_map = {
        'entry': 'reports/voucher_entry_pdf.html',
        'exit': 'reports/voucher_exit_pdf.html',
        'return': 'reports/voucher_return_pdf.html',
        'disposal': 'reports/voucher_disposal_pdf.html',
    }
    
    model = models_map.get(voucher_type)
    template = templates_map.get(voucher_type)
    
    if not model or not template:
        return HttpResponse("نوع الوصل غير صحيح", status=400)
    
    voucher = get_object_or_404(model, pk=pk)
    
    if not request.user.is_super_admin and voucher.tenant != request.user.tenant:
        return HttpResponse("ليس لديك صلاحية", status=403)
    
    items = voucher.items.select_related('product').prefetch_related('assets__inventory_item')
    
    html = render_to_string(template, {
        'voucher': voucher,
        'items': items,
        'tenant': request.user.tenant,
    })
    
    try:
        from weasyprint import HTML
        pdf = HTML(string=html).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{voucher.voucher_number}.pdf"'
        return response
    except ImportError:
        return HttpResponse("WeasyPrint غير مثبت", status=500)


from django.http import JsonResponse
